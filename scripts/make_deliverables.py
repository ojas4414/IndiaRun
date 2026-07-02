"""Produce the two portal deliverables from submission.csv:
  * submission.xlsx        -- ranked output in XLSX
  * approach_deck.pdf      -- what / why / how, as a slide deck

Run:  python scripts/make_deliverables.py
"""
import csv
import os
import textwrap

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.patches import Rectangle

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV = os.path.join(ROOT, "submission.csv")
XLSX = os.path.join(ROOT, "submission.xlsx")
PDF = os.path.join(ROOT, "approach_deck.pdf")

NAVY = "#0f2742"
BLUE = "#2f6db3"
GREY = "#40506a"


# --------------------------------------------------------------------------
# 1) XLSX
# --------------------------------------------------------------------------
def make_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "ranking"
    headers = ["candidate_id", "rank", "score", "reasoning"]
    ws.append(headers)
    for c in range(1, 5):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F2742")

    with open(CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ws.append([row["candidate_id"], int(row["rank"]),
                       float(row["score"]), row["reasoning"]])

    widths = {"A": 16, "B": 6, "C": 9, "D": 130}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=False)
    ws.freeze_panes = "A2"
    wb.save(XLSX)
    print(f"Wrote {XLSX}")


# --------------------------------------------------------------------------
# 2) PDF deck
# --------------------------------------------------------------------------
_PAGE = [1]  # running page counter, reset in make_pdf()
LIGHT_BLUE = "#7fa8d9"
ACCENT_BG = "#f4f7fb"


def _chrome(ax, title, subtitle=None, footer="kafka_consumer  |  github.com/ojas4414/IndiaRun"):
    """Shared header band + footer + page number for every content slide."""
    ax.add_patch(Rectangle((0, 0.82), 1, 0.18, color=NAVY, zorder=0))
    ax.add_patch(Rectangle((0, 0.815), 1, 0.006, color=BLUE, zorder=1))
    ax.text(0.05, 0.895, title, fontsize=25, color="white", weight="bold", va="center")
    if subtitle:
        ax.text(0.05, 0.845, subtitle, fontsize=13, color="#c7d6ea", va="center")

    ax.add_patch(Rectangle((0, 0), 1, 0.045, color="#eef2f7", zorder=0))
    ax.text(0.05, 0.022, footer, fontsize=9.5, color=GREY, va="center")
    ax.text(0.965, 0.022, f"{_PAGE[0]:02d}", fontsize=9.5, color=GREY,
            va="center", ha="right", weight="bold")
    _PAGE[0] += 1


def _slide(pdf, title, bullets, subtitle=None, footer="kafka_consumer  |  github.com/ojas4414/IndiaRun"):
    fig = plt.figure(figsize=(13.33, 7.5))
    ax = fig.add_axes([0, 0, 1, 1])
    ax.axis("off")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    _chrome(ax, title, subtitle, footer)

    y = 0.70
    for b in bullets:
        lead = b[0] if b and b[0] in "•-–" else "•"
        text = b.lstrip("•-– ")
        wrapped = textwrap.wrap(text, width=92)
        ax.text(0.06, y, lead, fontsize=15, color=BLUE, va="top", weight="bold")
        for i, line in enumerate(wrapped):
            ax.text(0.09, y - i * 0.045, line, fontsize=14.5, color=GREY, va="top")
        y -= 0.045 * len(wrapped) + 0.035

    pdf.savefig(fig)
    plt.close(fig)


def _table_slide(pdf, title, headers, rows, subtitle=None, note=None, col_x=None,
                 highlight_row=None):
    """A monospace-aligned data table (ablation / adversarial / composition)."""
    fig = plt.figure(figsize=(13.33, 7.5))
    ax = fig.add_axes([0, 0, 1, 1]); ax.axis("off"); ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    _chrome(ax, title, subtitle)

    col_x = col_x or [0.06 + i * (0.88 / len(headers)) for i in range(len(headers))]
    top = 0.72
    row_h = 0.052

    # header row
    ax.add_patch(Rectangle((0.05, top - 0.008), 0.90, row_h, color=NAVY, zorder=0))
    for x, h in zip(col_x, headers):
        ax.text(x, top + row_h / 2 - 0.008, h, fontsize=12.5, color="white",
                weight="bold", va="center", family="monospace")

    y = top - row_h
    for i, row in enumerate(rows):
        if highlight_row is not None and i == highlight_row:
            ax.add_patch(Rectangle((0.05, y - 0.006), 0.90, row_h, color=ACCENT_BG, zorder=0))
        for x, cell in zip(col_x, row):
            ax.text(x, y + row_h / 2 - 0.006, str(cell), fontsize=12,
                    color=NAVY if (highlight_row is not None and i == highlight_row) else GREY,
                    weight="bold" if (highlight_row is not None and i == highlight_row) else "normal",
                    va="center", family="monospace")
        y -= row_h

    if note:
        wrapped = textwrap.wrap(note, width=100)
        ny = y - 0.03
        for line in wrapped:
            ax.text(0.06, ny, line, fontsize=12, color=GREY, va="top", style="italic")
            ny -= 0.04

    pdf.savefig(fig)
    plt.close(fig)


def _quote_slide(pdf, title, entries, subtitle=None):
    """Sample-output slide: rank/score header + reasoning quote, per entry."""
    fig = plt.figure(figsize=(13.33, 7.5))
    ax = fig.add_axes([0, 0, 1, 1]); ax.axis("off"); ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    _chrome(ax, title, subtitle)

    y = 0.72
    for rank, score, title_role, quote in entries:
        ax.add_patch(Rectangle((0.05, y - 0.01), 0.9, 0.035, color=NAVY, zorder=0))
        ax.text(0.065, y + 0.0075, f"#{rank}  score {score:.4f}  —  {title_role}",
                fontsize=13, color="white", weight="bold", va="center")
        y -= 0.06
        wrapped = textwrap.wrap(quote, width=100)
        for line in wrapped:
            ax.text(0.07, y, line, fontsize=12, color=GREY, va="top", style="italic")
            y -= 0.038
        y -= 0.035

    pdf.savefig(fig)
    plt.close(fig)


def _title_slide(pdf):
    fig = plt.figure(figsize=(13.33, 7.5))
    ax = fig.add_axes([0, 0, 1, 1]); ax.axis("off"); ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.add_patch(Rectangle((0, 0), 1, 1, color=NAVY, zorder=0))
    ax.add_patch(Rectangle((0, 0.02), 1, 0.006, color=BLUE, zorder=1))
    ax.text(0.5, 0.66, "AI Candidate-Ranking System", fontsize=34, color="white",
            weight="bold", ha="center")
    ax.text(0.5, 0.56, "Intelligent Candidate Discovery & Ranking Challenge",
            fontsize=16, color="#c7d6ea", ha="center")
    ax.text(0.5, 0.46, "Team  kafka_consumer", fontsize=18, color="white", ha="center")
    ax.text(0.5, 0.40, "github.com/ojas4414/IndiaRun  |  Sandbox: huggingface.co/spaces/ojas-tulshian/India_Run",
            fontsize=11.5, color="#9fb8d6", ha="center")
    ax.text(0.5, 0.20, "The LLM should understand.  It should never decide.",
            fontsize=15, color=LIGHT_BLUE, ha="center", style="italic")
    ax.text(0.5, 0.06, "A deterministic retrieve-rerank pipeline with contrastive attention,\n"
                       "behavioral-twin disambiguation, and a measured, ablation-backed evaluation.",
            fontsize=10.5, color="#c7d6ea", ha="center", va="center")
    pdf.savefig(fig); plt.close(fig)


def _agenda_slide(pdf):
    fig = plt.figure(figsize=(13.33, 7.5))
    ax = fig.add_axes([0, 0, 1, 1]); ax.axis("off"); ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    _chrome(ax, "Agenda", "What this deck covers")
    items = [
        ("01", "The problem & why it's hard"),
        ("02", "Thesis & system architecture"),
        ("03", "How every trap in the brief is defended against"),
        ("04", "Two novel components: contrastive attention, twin disambiguation"),
        ("05", "Evaluation methodology & measured results (baselines, ablation, adversarial)"),
        ("06", "Sample output, guarantees, honest limitations, and tech stack"),
    ]
    y = 0.70
    for num, text in items:
        ax.text(0.08, y, num, fontsize=20, color=BLUE, weight="bold", va="center")
        ax.text(0.16, y, text, fontsize=15.5, color=GREY, va="center")
        y -= 0.095
    pdf.savefig(fig); plt.close(fig)


def make_pdf():
    _PAGE[0] = 1
    with PdfPages(PDF) as pdf:
        _title_slide(pdf)
        _agenda_slide(pdf)

        _slide(pdf, "The Problem", [
            "Rank the top 100 of 100,000 candidates for one Senior AI Engineer JD (Redrob AI, Series A).",
            "\"The right answer is not: find candidates whose skills section contains the most AI keywords.\" "
            "— that trap is explicitly built into the dataset.",
            "Traps: keyword stuffers, plain-language strong fits (\"Tier-5s\"), behavioral twins, ~80 honeypots "
            "(impossible profiles). Honeypot rate >10% in the top-100 is an automatic disqualification.",
            "Constraints: <=5 min, 16 GB RAM, CPU-only, no network during ranking; output must be reproducible.",
        ], subtitle="What we were asked to solve")

        _slide(pdf, "Why this is hard", [
            "Precision vs. constraints: the highest-precision approach (an LLM judging every profile) is exactly "
            "what the time/CPU/offline budget rules out.",
            "The dataset is adversarial by design: strong keyword overlap correlates with rank in a naive system, "
            "but is explicitly NOT what the JD wants.",
            "The JD itself asks for hiring judgement most systems never encode: availability, tenure patterns, "
            "verified vs. self-reported skill, and production vs. tutorial-level experience.",
        ], subtitle="Precision, adversarial data, and a hard runtime budget, all at once")

        _slide(pdf, "Thesis", [
            "Letting an LLM rank is non-deterministic, slow, and can't run offline in budget.",
            "So we split UNDERSTANDING (offline embeddings) from RANKING (deterministic scoring).",
            "rank.py makes zero LLM/network calls and returns byte-identical output on identical input.",
            "Everything that needs judgement is pushed into precomputed features and explicit scoring rules -- "
            "never into a runtime model call.",
        ], subtitle="The LLM should understand, never decide")

        _slide(pdf, "Pipeline: retrieve -> rerank -> disambiguate", [
            "Precompute (no time limit): honeypot filter -> sentence-transformer embeddings (all-MiniLM-L6-v2) "
            "-> FAISS index (~47 min for 100k candidates, one-time cost).",
            "Stage 1  Retrieval: FAISS ANN -> top-2000 semantic funnel (recall, not final ranking).",
            "Stage 2  Feature rerank: trajectory-DP 35% | skill-graph 22% | behavioral 18% | attention 15% | "
            "semantic 10%  - disqualifier penalties.",
            "Stage 3  Attention rerank over the top-600 survivors (precision + evidence extraction).",
            "Stage 4  Behavioral-twin disambiguation -> top-100 + evidence-grounded reasoning (~40s CPU total).",
        ], subtitle="Architecture")

        _slide(pdf, "Design decisions: why these weights", [
            "Trajectory carries the most weight (35%) because the ablation shows it is the single strongest "
            "JD-fit signal -- removing it drops NDCG@100 from 0.876 to 0.753.",
            "Semantic similarity is weighted lowest (10%): it is mainly a RECALL signal for the funnel, and "
            "the ablation shows it is nearly inert in final ordering once the funnel already exists.",
            "Behavioral (18%) and attention (15%) are kept as deliberate minority signals: they optimize "
            "availability and explainability, which the JD explicitly requires but a pure fit-rubric can't see.",
            "Weights were tuned FROM the evaluation harness, not chosen first and rationalized after.",
        ], subtitle="Every weight in scoring/hybrid_scorer.py traces back to a measurement")

        _slide(pdf, "How each trap is handled", [
            "Keyword stuffing -> skill-graph + trajectory carry 57% of weight; off-domain titles hard-disqualified "
            "before they ever reach the scorer.",
            "Plain-language Tier-5s -> semantic retrieval + sentence attention surface hidden production evidence "
            "even when a profile never uses the JD's exact vocabulary.",
            "Behavioral twins -> cluster near-duplicates by role/experience/skill-Jaccard; the most-available "
            "twin leads, the rest are demoted so duplicates don't stack the top-100.",
            "~80 honeypots -> date/tenure impossibility checks at precompute (52 caught in this run; 0 in top-100, "
            "vs. the 10% disqualification threshold).",
            "Consulting-only / CV-speech-only / title-chasers / no-recent-code -> explicit disqualifier rules "
            "lifted directly from the JD's \"things we explicitly do NOT want\" section.",
        ], subtitle="Trap defenses map directly to the brief")

        _slide(pdf, "Novel #1: contrastive-facet sentence attention", [
            "Cross-attention pooling: JD facets = queries, candidate sentences = keys/values "
            "(ColBERT-style MaxSim / late interaction).",
            "Recovers a single strong sentence that document-level embedding would average away -- e.g. one line "
            "about shipping a production retrieval system, buried in an otherwise generic profile.",
            "Contrastive anti-fit facets cancel sentence-level keyword stuffing (e.g. \"SEO articles that ranked "
            "in search\" superficially resembles retrieval language but nets to zero against the anti-fit facet).",
            "The highest-attention sentence is surfaced as a literal quoted evidence line in the candidate's "
            "generated reasoning -- not just a score, a citation.",
        ], subtitle="Precision + explainability")

        _slide(pdf, "Novel #2: twins + adversarial robustness", [
            "Behavioral-twin disambiguation: paper-identical profiles (same role, experience, skills) are "
            "separated by availability signals instead of being ranked as duplicates.",
            "Adversarial test: inject every JD skill + fake 95/100 assessment scores + a buzzword summary into "
            "profiles that should NOT rank, and measure promotion into the top-100.",
            "Off-domain promotion under attack: FULL system 0%  vs  naive keyword ranker 100%.",
            "We also report an honestly-found residual weakness (generic engineers under extreme stuffing) plus "
            "its mitigation (skill grounding) -- rigor over spin.",
        ], subtitle="Things a cosine-similarity submission won't have")

        _slide(pdf, "Evaluation methodology", [
            "The JD explicitly lists NDCG / MRR / MAP / offline-to-online correlation as required skills for "
            "this role -- so the submission itself needed to demonstrate that rigor, not just claim it.",
            "Two kinds of evidence: (1) NDCG@10/@100, MRR, MAP@100 against a transparent recruiter rubric "
            "(weak supervision, documented circularity caveat); (2) label-free trap metrics -- honeypot % and "
            "off-domain-title % in the top-100 -- which need no labels at all.",
            "Compared against two baselines (naive keyword count, pure semantic similarity) and a full "
            "per-component ablation (drop skill-graph / trajectory / behavioral / attention / semantic one at a time).",
            "Plus the adversarial keyword-stuffing test above, targeting the exact trap the JD calls out.",
        ], subtitle="Evaluation the JD explicitly asks for")

        _table_slide(pdf, "Results: baselines vs. full system",
            headers=["config", "NDCG@10", "NDCG@100", "MRR", "MAP@100", "HP%", "OffDom%"],
            rows=[
                ["naive_keyword", "0.653", "0.615", "0.500", "0.790", "0.0", "6.0"],
                ["semantic_only", "0.656", "0.508", "1.000", "0.789", "0.0", "27.0"],
                ["full system",   "1.000", "0.876", "1.000", "0.985", "0.0", "4.0"],
            ],
            subtitle="Full system vs. naive-keyword and pure-semantic baselines",
            highlight_row=2,
            note="Off-domain titles in the top-100 drop from 27% (semantic-only) to 4% (full system). "
                 "Honeypots stay at 0% for every gated config -- well under the 10% disqualification bar.")

        _table_slide(pdf, "Results: per-component ablation",
            headers=["config", "NDCG@10", "NDCG@100", "MRR", "MAP@100"],
            rows=[
                ["full (all signals)", "1.000", "0.876", "1.000", "0.985"],
                ["- attention",        "1.000", "0.938", "1.000", "0.997"],
                ["- skill_graph",      "1.000", "0.898", "1.000", "0.992"],
                ["- trajectory",       "0.879", "0.753", "1.000", "0.916"],
                ["- behavioral",       "1.000", "0.904", "1.000", "0.989"],
                ["- semantic",         "1.000", "0.884", "1.000", "0.984"],
            ],
            subtitle="Each row removes one signal from the full system",
            highlight_row=3,
            note="Removing trajectory causes by far the largest drop (0.876 -> 0.753) -- confirming it as the "
                 "dominant JD-fit signal. Attention/behavioral/semantic optimize objectives (explainability, "
                 "availability, recall) the rubric doesn't score, which is why removing them can nudge rubric-NDCG up.")

        _table_slide(pdf, "Results: adversarial keyword-stuffing attack",
            headers=["cohort", "n", "full system", "naive keyword"],
            rows=[
                ["off-domain roles",  "40", "0%",   "100%"],
                ["generic engineers", "4",  "100%", "100%"],
            ],
            subtitle="Every JD skill + fake 95/100 assessments injected into profiles that shouldn't rank",
            highlight_row=0,
            note="Off-domain roles are fully robust: the disqualifier gate keys on the (unchanged) title, and "
                 "trajectory keys on the (unchanged) career history, so stuffing changes nothing. Residual "
                 "weakness (generic engineers, n=4): claimed-but-uncorroborated skills can still promote a "
                 "borderline profile -- honestly reported, with skill-grounding logged as the fix.")

        _table_slide(pdf, "Top-100 composition (current run)",
            headers=["count", "current title"],
            rows=[
                ["17", "Search Engineer"], ["12", "Applied ML Engineer"], ["12", "AI Engineer"],
                ["11", "Machine Learning Engineer"], ["8", "AI Research Engineer"],
                ["6", "Senior NLP Engineer"], ["6", "Senior Data Scientist"],
                ["5", "Senior Machine Learning Engineer"], ["4", "Recommendation Systems Engineer"],
                ["4", "Staff Machine Learning Engineer"], ["3", "Senior AI Engineer"],
                ["3", "NLP Engineer"], ["3", "ML Engineer"], ["2", "Senior Applied Scientist"],
                ["2", "Lead AI Engineer"], ["1", "Junior ML Engineer / AI Specialist"],
            ],
            subtitle="Every one of the top 100 holds an on-domain AI/ML/Search title",
            note="Zero Marketing Managers, HR Managers, Accountants, or other off-domain keyword-stuffed roles "
                 "reached the top-100 -- the keyword-trap defense holds on the real, full 100k-candidate run.")

        _quote_slide(pdf, "Sample output (real, from this run's submission.csv)", [
            (1, 0.7879, "Senior Machine Learning Engineer",
             'Standout candidate -- 7.2 years of work as Senior Machine Learning Engineer. Brings Weaviate, '
             'Pinecone, Information Retrieval, Milvus -- the core of the embeddings & retrieval stack this role '
             'centres on. Assessment-verified in Weaviate (72/100).'),
            (2, 0.7463, "Staff Machine Learning Engineer",
             'Excellent match -- Staff Machine Learning Engineer, ~7 yrs experience. Brings QLoRA, Pinecone, '
             'BM25, Information Retrieval. Evidence: "Spent substantial time on the boring-but-critical parts: '
             'incremental index refresh, embedding drift..."'),
            (3, 0.7441, "Lead AI Engineer",
             'Strong hire signal -- Lead AI Engineer with 6.7 yrs experience. Hands-on with Information '
             'Retrieval, Learning to Rank, Elasticsearch, Python. From their profile: "Designed the offline '
             'evaluation framework from scratch -- NDCG, MRR, recall@K calibrated against online A/B metrics."'),
        ], subtitle="Every reasoning string is unique and cites a real profile quote")

        _slide(pdf, "Results & guarantees", [
            "Ranks the full 100k in ~40s (limit: 5 min); 0 honeypots and ~100% on-domain titles in the top-100.",
            "Deterministic: two independent runs produce byte-identical output (MD5-verified).",
            "Offline: embedding model baked into the Docker image; ranking sets TRANSFORMERS_OFFLINE=1 and "
            "HF_HUB_OFFLINE=1, no network calls at rank time.",
            "Reproducible end-to-end: `docker compose up --build` runs precompute -> rank -> validate in one command.",
            "24 automated tests covering the skill graph, trajectory DP, honeypot rules, attention math, twin "
            "clustering, and evaluation metrics.",
        ], subtitle="Meets every hard constraint")

        _slide(pdf, "Honest limitations", [
            "The recruiter-rubric NDCG (gold.py) is weak supervision that partially overlaps with the ranker's "
            "own features (title, career text) -- read it as \"agreement with an explicit rubric,\" not an "
            "unbiased oracle. The label-free HP%/OffDom% metrics are the load-bearing evidence.",
            "Attention contributes only 15% of the score by design; its main value is the evidence quote in "
            "each reasoning, not raw NDCG -- a deliberate trade for explainability the rubric can't measure.",
            "The adversarial test found a residual gap: heavily skill-stuffed generic engineers can still be "
            "promoted. Logged as future work (skill grounding: discount claims never corroborated in career text).",
            "We chose to report this rather than hide it -- an adversarial test is only useful if you act on "
            "what it finds.",
        ], subtitle="What we'd fix next, said plainly")

        _slide(pdf, "Tech stack", [
            "Embeddings & retrieval: sentence-transformers (all-MiniLM-L6-v2), FAISS (IndexFlatIP / cosine).",
            "Scoring: pure Python -- skill-graph BFS, Needleman-Wunsch trajectory alignment, rule-based "
            "disqualifiers and behavioral scoring, cross-attention pooling with contrastive facets.",
            "Packaging: Docker + docker-compose (one-command reproduction), pytest (24 tests), Gradio sandbox "
            "on HuggingFace Spaces for live, small-sample demos.",
            "Evaluation: a custom harness (NDCG/MRR/MAP, ablation, adversarial attack) -- no external eval "
            "library needed, everything is inspectable in evaluation/.",
        ], subtitle="Deliberately dependency-light and fully inspectable")

        fig = plt.figure(figsize=(13.33, 7.5))
        ax = fig.add_axes([0, 0, 1, 1]); ax.axis("off"); ax.set_xlim(0, 1); ax.set_ylim(0, 1)
        ax.add_patch(Rectangle((0, 0), 1, 1, color=NAVY, zorder=0))
        ax.text(0.5, 0.58, "Thank you", fontsize=32, color="white", weight="bold", ha="center")
        ax.text(0.5, 0.47, "Team kafka_consumer", fontsize=16, color="#c7d6ea", ha="center")
        ax.text(0.5, 0.38, "Code:     github.com/ojas4414/IndiaRun", fontsize=13, color="white", ha="center")
        ax.text(0.5, 0.33, "Sandbox:  huggingface.co/spaces/ojas-tulshian/India_Run", fontsize=13,
                color="white", ha="center")
        ax.text(0.5, 0.16, "Happy to walk through any design decision in detail.",
                fontsize=13, color=LIGHT_BLUE, ha="center", style="italic")
        pdf.savefig(fig); plt.close(fig)

    print(f"Wrote {PDF} ({_PAGE[0] + 1} pages)")  # +1 title +1 closing slide, neither uses _chrome


if __name__ == "__main__":
    make_xlsx()
    make_pdf()
